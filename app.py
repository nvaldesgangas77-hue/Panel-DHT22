from flask import Flask, render_template, jsonify, request, redirect, url_for, session, send_file
import serial, time, os, csv, pandas as pd
import io
import matplotlib.pyplot as plt
from fpdf import FPDF
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import threading
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv


load_dotenv()

import requests

# ---------------- API Meteorológica (OpenWeather) ----------------
def obtener_clima():
    """Consulta OpenWeatherMap y retorna temperatura y humedad externas."""
    try:
        API_KEY = os.getenv("OWM_API_KEY")
        LAT, LON = -37.47, -72.35  # 📍 Coordenadas de Los Ángeles, Biobío (puedes cambiar)
        url = f"https://api.openweathermap.org/data/2.5/weather?lat={LAT}&lon={LON}&appid={API_KEY}&units=metric&lang=es"
        r = requests.get(url, timeout=5)
        data = r.json()
        temp_ext = data["main"]["temp"]
        hum_ext = data["main"]["humidity"]
        cond = data["weather"][0]["description"]
        print(f"🌤️ Clima externo → {temp_ext} °C | {hum_ext}% | {cond}")
        return temp_ext, hum_ext, cond
    except Exception as e:
        print(f"⚠️ Error al obtener clima externo: {e}")
        return None, None, None



# ---------------- Configuración ----------------
PORT = "COM3"
BAUD = 9600
arduino = None

app = Flask(__name__)
app.secret_key = "supersecreto_nico"
app.permanent_session_lifetime = timedelta(minutes=15)
# ---------------- Umbrales de temperatura y humedad ----------------
TEMP_MIN = 30.0
TEMP_MAX = 39.0
HUM_MIN = 40.0
HUM_MAX = 85.0

alerta_activa = {"temp": False, "hum": False, "msg": ""}


# ---------------- Fechas comparadas ----------------
fechas_comparadas = []  # Guarda las fechas seleccionadas desde el panel web

@app.route("/set_fechas_comparadas", methods=["POST"])
def set_fechas_comparadas():
    """Recibe desde el navegador las fechas que el usuario está comparando."""
    global fechas_comparadas
    try:
        data = request.get_json(force=True)
        fechas_comparadas = data.get("fechas", [])
        print(f"📅 Fechas comparadas registradas: {fechas_comparadas}")
        return jsonify({"status": "ok", "fechas": fechas_comparadas})
    except Exception as e:
        print(f"Error al registrar fechas comparadas: {e}")
        return jsonify({"status": "error", "detalle": str(e)})


# ---------------- Rutas protegidas ----------------
def login_requerido(func):
    def wrapper(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

# ---------------- Base de datos usuarios ----------------
DB_USUARIOS = "usuarios.db"

def init_db():
    conn = sqlite3.connect(DB_USUARIOS)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS usuarios (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    usuario TEXT UNIQUE,
                    contrasena TEXT
                )""")
    conn.commit()
    c.execute("SELECT * FROM usuarios WHERE usuario=?", ("admin",))
    if not c.fetchone():
        c.execute("INSERT INTO usuarios (usuario, contrasena) VALUES (?, ?)",
                  ("admin", generate_password_hash("1234")))
        conn.commit()
        print("✅ Usuario admin creado (usuario: admin, contraseña: 1234)")
    conn.close()

init_db()

# ---------------- Conexión Arduino ----------------
def conectar_arduino():
    global arduino
    if arduino and arduino.is_open:
        return arduino
    try:
        arduino = serial.Serial(PORT, BAUD, timeout=2)
        time.sleep(2)
        print(f"✅ Conectado correctamente al puerto {PORT}")
        return arduino
    except Exception as e:
        print(f"⚠️ No se pudo conectar con Arduino: {e}")
        return None

# ---------------- Variables globales ----------------
ultima_temp = 0.0
ultima_hum = 0.0
buffer_temp, buffer_hum = [], []
ultimo_promedio = datetime.now()

def guardar_datos(temp, hum):
    global buffer_temp, buffer_hum, ultimo_promedio

    fecha_actual = datetime.now()
    año = fecha_actual.strftime("%Y")
    mes = fecha_actual.strftime("%m_%B")
    dia = fecha_actual.strftime("%Y-%m-%d")

    # 📁 Crear carpeta donde guardar promedios
    carpeta_mes = os.path.join("data", año, mes)
    os.makedirs(carpeta_mes, exist_ok=True)

    # 🔹 Obtener clima externo (solo una vez por lectura)
    temp_ext, hum_ext, cond = obtener_clima()

    # --- Acumular lecturas ---
    buffer_temp.append(temp)
    buffer_hum.append(hum)

    # --- Evaluar si hay alerta inmediata (solo notificación, no guardado) ---
    global alerta_activa
    alerta_msg = ""
    if temp < TEMP_MIN or temp > TEMP_MAX:
        alerta_activa["temp"] = True
        alerta_msg += f"⚠️ Temperatura fuera de rango: {temp:.1f} °C (ideal 30–38 °C). "
    else:
        alerta_activa["temp"] = False

    if hum < HUM_MIN or hum > HUM_MAX:
        alerta_activa["hum"] = True
        alerta_msg += f"⚠️ Humedad fuera de rango: {hum:.1f}% (ideal 50–80%). "
    else:
        alerta_activa["hum"] = False

    alerta_activa["msg"] = alerta_msg.strip()
    if alerta_msg:
        enviar_correo("[HiveSim] ⚠️ Alerta ambiental en colmena", alerta_msg)
        print(alerta_msg)

    # --- Cada 10 lecturas (≈10 min) guardar promedio ---
    if len(buffer_temp) >= 10 or (datetime.now() - ultimo_promedio).total_seconds() >= 600:
        prom_t = sum(buffer_temp) / len(buffer_temp)
        prom_h = sum(buffer_hum) / len(buffer_hum)
        temp_ext, hum_ext, cond = obtener_clima()

        # --- Guardar con formato en Excel (.xlsx) ---
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # Crear o cargar archivo Excel
        archivo_prom = os.path.join(carpeta_mes, f"{dia}_promedios.xlsx")
        existe_prom = os.path.exists(archivo_prom)

        if existe_prom:
            wb = load_workbook(archivo_prom)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Promedios"

            # Encabezados
            headers = [
                "Fecha", "Hora",
                "Temp promedio (°C)", "Hum promedio (%)",
                "Temp externa (°C)", "Hum externa (%)",
                "Condición climática", "Alerta temperatura",
                "Alerta humedad", "Mensaje"
            ]
            ws.append(headers)

            # Estilo para encabezados
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin = Side(border_style="thin", color="000000")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # --- Evaluar alertas con doble umbral ---
        alerta_temp = "Óptima"
        alerta_hum = "Óptima"
        mensaje = ""

        # Temperatura
        if prom_t < 30 or prom_t > 38:
            alerta_temp = "Crítica"
            mensaje += f"🚨 Temperatura crítica ({prom_t:.1f} °C; fuera del rango 30–38 °C). "
        elif prom_t < 32 or prom_t > 36:
            alerta_temp = "Precaución"
            mensaje += f"⚠️ Temperatura fuera del rango óptimo ({prom_t:.1f} °C; ideal 32–36 °C). "

        # Humedad
        if prom_h < 40 or prom_h > 85:
            alerta_hum = "Crítica"
            mensaje += f"🚨 Humedad crítica ({prom_h:.1f}% fuera del rango 40–85%). "
        elif prom_h < 50 or prom_h > 75:
            alerta_hum = "Precaución"
            mensaje += f"⚠️ Humedad fuera del rango óptimo ({prom_h:.1f}%; ideal 50–75%). "

        # Agregar fila
        nueva_fila = [
            dia, fecha_actual.strftime("%H:%M:%S"),
            round(prom_t, 2), round(prom_h, 2),
            temp_ext if temp_ext is not None else "N/A",
            hum_ext if hum_ext is not None else "N/A",
            cond if cond else "N/A",
            alerta_temp, alerta_hum, mensaje.strip()
        ]
        ws.append(nueva_fila)

        # Aplicar estilo general de bordes
        thin = Side(border_style="thin", color="000000")
        for col in ws.columns:
            for cell in col:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # --- Colorear filas según el nivel de alerta ---
        # Verde = Óptima | Amarillo = Precaución | Rojo = Crítica

        fill_optima = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde suave
        fill_precaucion = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Amarillo
        fill_critica = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Rojo claro

        # Determinar color
        if "Crítica" in (alerta_temp, alerta_hum):
            fill = fill_critica
        elif "Precaución" in (alerta_temp, alerta_hum):
            fill = fill_precaucion
        else:
            fill = fill_optima

        # Aplicar color a la última fila agregada
        for cell in ws[ws.max_row]:
            cell.fill = fill

        # Ajustar ancho automático
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Guardar y cerrar archivo
        wb.save(archivo_prom)
        wb.close()

        buffer_temp.clear()
        buffer_hum.clear()
        ultimo_promedio = datetime.now()
        print(f"🟢 Promedio guardado → 🌡️ {prom_t:.2f} °C | 💧 {prom_h:.2f} %")



# ---------------- Leer datos Arduino ----------------
def leer_datos():
    global ultima_temp, ultima_hum, arduino
    if not arduino or not arduino.is_open:
        arduino = conectar_arduino()
        return ultima_temp, ultima_hum

    try:
        linea = arduino.readline().decode(errors="ignore").strip()
        if linea:
            partes = linea.split(",")
            if len(partes) == 2:
                temp = float(partes[0])
                hum = float(partes[1])
                ultima_temp, ultima_hum = temp, hum
                guardar_datos(temp, hum)
                print(f"🌡️ {temp:.2f} °C | 💧 {hum:.2f} %")
                return temp, hum
    except Exception as e:
        print(f"⚠️ Error leyendo datos: {e}")
    return ultima_temp, ultima_hum

# ---------------- Funciones auxiliares ----------------
def limpiar_columnas(df):
    posibles_temp = [c for c in df.columns if "Temp" in c]
    posibles_hum = [c for c in df.columns if "Hum" in c]
    if posibles_temp and posibles_hum:
        df["Temperatura (°C)"] = pd.to_numeric(df[posibles_temp[0]], errors="coerce")
        df["Humedad (%)"] = pd.to_numeric(df[posibles_hum[0]], errors="coerce")
    return df.dropna(subset=["Temperatura (°C)", "Humedad (%)"])

# ---------------- Rutas principales ----------------
@app.route("/")
@login_requerido
def index():
    return render_template("index.html", usuario=session["usuario"])

@app.route("/data")
@login_requerido
def data():
    global buffer_temp, buffer_hum, ultimo_promedio

    if len(buffer_temp) >= 10:
        prom_t = sum(buffer_temp) / len(buffer_temp)
        prom_h = sum(buffer_hum) / len(buffer_hum)
        buffer_temp.clear()
        buffer_hum.clear()
        ultimo_promedio = datetime.now()
        print(f"📈 Promedio enviado → 🌡️ {prom_t:.2f} °C | 💧 {prom_h:.2f} %")
        return jsonify({"temperatura": prom_t, "humedad": prom_h})

    return jsonify({"temperatura": ultima_temp, "humedad": ultima_hum})

@app.route("/api/data", methods=["POST"])
def recibir_datos():
    global ultima_temp, ultima_hum
    try:
        data = request.get_json(force=True)
        temp = float(data.get("temperatura", 0))
        hum = float(data.get("humedad", 0))
        ultima_temp, ultima_hum = temp, hum
        guardar_datos(temp, hum)
        print(f"📡 Dato recibido Wi-Fi → 🌡️ {temp:.2f} °C | 💧 {hum:.2f} %")
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        print(f"❌ Error recibiendo datos Wi-Fi: {e}")
        return jsonify({"status": "error", "detalle": str(e)}), 400
# ---------------- Estado de alerta para el panel ----------------
@app.route("/alerta")
@login_requerido
def alerta():
    return jsonify(alerta_activa)

# ---------------- Histórico ----------------
@app.route("/historico_data")
@login_requerido
def historico_data():
    return procesar_historico_por_fecha(datetime.now().strftime("%Y-%m-%d"))

@app.route("/historico_por_fecha/<fecha>")
@login_requerido
def historico_por_fecha(fecha):
    return procesar_historico_por_fecha(fecha)

def procesar_historico_por_fecha(fecha):
    try:
        ruta = None
        for root, _, files in os.walk("data"):
            for f in files:
                if f.endswith(f"{fecha}_promedios.csv"):
                    ruta = os.path.join(root, f)
                    break

        if not ruta:
            return jsonify({"error": "No hay datos disponibles"})

        df = None  # inicializamos antes del try

        # Intentar con diferentes codificaciones y saltar líneas corruptas
        for cod in ["utf-8", "latin1", "iso-8859-1"]:
            try:
                df = pd.read_csv(ruta, encoding=cod, on_bad_lines="skip")
                break
            except Exception:
                continue

        if df is None or df.empty:
            return jsonify({"error": "No se pudieron leer los datos del archivo"})

        # Normalizar nombres de columnas si vienen con distinto formato
        if "Temp promedio (°C)" in df.columns and "Hum promedio (%)" in df.columns:
            df.rename(columns={
                "Temp promedio (°C)": "Temperatura (°C)",
                "Hum promedio (%)": "Humedad (%)"
            }, inplace=True)

        df = limpiar_columnas(df)

        if df.empty:
            return jsonify({"error": "No hay datos válidos para graficar"})

        # Validar existencia de columnas
        if "Temperatura (°C)" not in df.columns or "Humedad (%)" not in df.columns:
            return jsonify({"error": "Columnas inválidas en el CSV"})

        # Procesar y limpiar formato de hora
        df["Hora"] = pd.to_datetime(df["Hora"], errors="coerce")
        df = df.dropna(subset=["Hora"])
        df["Hora_str"] = df["Hora"].dt.strftime("%H:%M")

        # Agrupar promedios
        temp_prom = df.groupby("Hora_str")["Temperatura (°C)"].mean().round(2).tolist()
        hum_prom = df.groupby("Hora_str")["Humedad (%)"].mean().round(2).tolist()
        horas = df["Hora_str"].unique().tolist()

        if not horas or not temp_prom or not hum_prom:
            return jsonify({"error": "Sin datos válidos para graficar"})

        return jsonify({
            "fechas": horas,
            "temp_mean": temp_prom,
            "hum_mean": hum_prom
        })

    except Exception as e:
        print(f"❌ Error procesando histórico: {e}")
        return jsonify({"error": str(e)})


        for cod in ["utf-8", "latin1", "iso-8859-1"]:
            try:
                df = pd.read_csv(ruta, encoding=cod)
                break
            except Exception:
                continue

        if "Temp promedio (°C)" in df.columns and "Hum promedio (%)" in df.columns:
            df.rename(columns={
                "Temp promedio (°C)": "Temperatura (°C)",
                "Hum promedio (%)": "Humedad (%)"
            }, inplace=True)

        df = limpiar_columnas(df)
        if df.empty:
            return jsonify({"error": "No hay datos válidos"})

        df["Hora"] = pd.to_datetime(df["Hora"], errors="coerce")
        df = df.dropna(subset=["Hora"])
        df["Hora_str"] = df["Hora"].dt.strftime("%H:%M")

        temp_prom = df.groupby("Hora_str")["Temperatura (°C)"].mean().round(2).tolist()
        hum_prom = df.groupby("Hora_str")["Humedad (%)"].mean().round(2).tolist()
        horas = df["Hora_str"].unique().tolist()

        return jsonify({
            "fechas": horas,
            "temp_mean": temp_prom,
            "hum_mean": hum_prom
        })
    except Exception as e:
        print(f"❌ Error procesando histórico: {e}")
        return jsonify({"error": str(e)})

# ---------------- Generación de Informe PDF ----------------
@app.route("/generar_pdf")
@login_requerido
def generar_pdf():
    try:
        hoy = datetime.now().strftime("%Y-%m-%d")
        ruta_hoy = None
        ruta_cmp = None

        # Buscar CSV del día actual
        for root, _, files in os.walk("data"):
            for f in files:
                if f.endswith(f"{hoy}_promedios.csv"):
                    ruta_hoy = os.path.join(root, f)
                elif "_promedios.csv" in f and f != f"{hoy}_promedios.csv":
                    ruta_cmp = os.path.join(root, f)

        if not ruta_hoy:
            return jsonify({"error": "No hay datos de hoy para generar el informe"})

        # --- Leer CSV actual con tolerancia de codificación ---
        try:
            df_hoy = pd.read_csv(ruta_hoy, encoding="utf-8")
        except UnicodeDecodeError:
            df_hoy = pd.read_csv(ruta_hoy, encoding="latin1")

        if "Temp promedio (°C)" in df_hoy.columns:
            df_hoy.rename(columns={
                "Temp promedio (°C)": "Temperatura (°C)",
                "Hum promedio (%)": "Humedad (%)"
            }, inplace=True)

        # --- Leer CSV comparado solo si existe ---
        df_cmp = None
        if ruta_cmp:
            try:
                df_cmp = pd.read_csv(ruta_cmp, encoding="utf-8")
            except UnicodeDecodeError:
                df_cmp = pd.read_csv(ruta_cmp, encoding="latin1")

            if "Temp promedio (°C)" in df_cmp.columns:
                df_cmp.rename(columns={
                    "Temp promedio (°C)": "Temperatura (°C)",
                    "Hum promedio (%)": "Humedad (%)"
                }, inplace=True)

        # --- Gráfico comparativo mejorado ---
        plt.figure(figsize=(8, 4))

        # Convertir hora a formato legible y asegurar orden
        df_hoy["Hora"] = pd.to_datetime(df_hoy["Hora"], errors="coerce")
        df_hoy = df_hoy.dropna(subset=["Hora"]).sort_values("Hora")

        plt.plot(df_hoy["Hora"], df_hoy["Temperatura (°C)"], color="#ff6b6b", label="Temperatura hoy (°C)")
        plt.plot(df_hoy["Hora"], df_hoy["Humedad (%)"], color="#00b4d8", label="Humedad hoy (%)")

        if df_cmp is not None:
            df_cmp["Hora"] = pd.to_datetime(df_cmp["Hora"], errors="coerce")
            df_cmp = df_cmp.dropna(subset=["Hora"]).sort_values("Hora")

            plt.plot(df_cmp["Hora"], df_cmp["Temperatura (°C)"], color="#ffb703", linestyle="--",
                     label="Temperatura comparada (°C)")
            plt.plot(df_cmp["Hora"], df_cmp["Humedad (%)"], color="#fb8500", linestyle="--",
                     label="Humedad comparada (%)")

        plt.xlabel("Hora")
        plt.ylabel("Valor")
        plt.title(f"Comparativa de Temperatura y Humedad ({hoy})")
        plt.legend()
        plt.grid(alpha=0.3)

        # Rotar etiquetas y reducir densidad
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        img_buf = io.BytesIO()
        plt.savefig(img_buf, format="png", dpi=120, bbox_inches="tight")
        plt.close()
        img_buf.seek(0)

        # --- Datos resumen ---
        temp_prom = df_hoy["Temperatura (°C)"].mean()
        hum_prom = df_hoy["Humedad (%)"].mean()
        temp_max = df_hoy["Temperatura (°C)"].max()
        temp_min = df_hoy["Temperatura (°C)"].min()
        hum_max = df_hoy["Humedad (%)"].max()
        hum_min = df_hoy["Humedad (%)"].min()

        # --- Crear PDF ---
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 18)
        pdf.set_text_color(0, 100, 200)
        pdf.cell(0, 10, "HiveSim - Informe Diario de Colmena", ln=True, align="C")
        pdf.set_font("Arial", "", 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, f"Fecha: {hoy}", ln=True, align="C")
        pdf.ln(8)

        pdf.set_font("Arial", "B", 14)
        pdf.set_text_color(0, 100, 200)
        pdf.cell(0, 10, "Resumen General", ln=True)
        pdf.set_font("Arial", "", 12)
        pdf.set_text_color(0, 0, 0)

        pdf.cell(0, 8, f"Temperatura promedio: {temp_prom:.2f} °C", ln=True)
        pdf.cell(0, 8, f"Humedad promedio: {hum_prom:.2f} %", ln=True)
        pdf.cell(0, 8, f"Temperatura máxima: {temp_max:.2f} °C | mínima: {temp_min:.2f} °C", ln=True)
        pdf.cell(0, 8, f"Humedad máxima: {hum_max:.2f} % | mínima: {hum_min:.2f} %", ln=True)
        pdf.ln(6)

        # --- Comentario automático ---
        estado = "Estable"
        comentario = "El sistema muestra condiciones normales dentro del rango térmico y de humedad adecuado."

        if temp_max > 36 or temp_min < 32:
            estado = "Temperatura fuera de rango"
            comentario = f"Temperatura fuera de rango detectada (mín: {temp_min:.2f} °C, máx: {temp_max:.2f} °C)."
        if hum_max > 80 or hum_min < 50:
            estado = "Humedad fuera de rango"
            comentario += f" Humedad fuera de rango detectada (mín: {hum_min:.2f} %, máx: {hum_max:.2f} %)."

        pdf.set_font("Arial", "B", 14)
        pdf.set_text_color(0, 100, 200)
        pdf.cell(0, 10, "Análisis del Día", ln=True)
        pdf.set_font("Arial", "", 12)
        pdf.set_text_color(0, 0, 0)
        pdf.multi_cell(0, 8, comentario)
        pdf.ln(10)

        # --- Insertar gráfico ---
        img_path = "grafico_temp_hum.png"
        with open(img_path, "wb") as f:
            f.write(img_buf.getvalue())
        pdf.image(img_path, x=20, w=170)
        os.remove(img_path)

        pdf.ln(10)

        # --- Referencias científicas ---
        pdf.set_font("Arial", "B", 14)
        pdf.set_text_color(0, 100, 200)
        pdf.cell(0, 10, "Referencias", ln=True)
        pdf.set_font("Arial", "", 10)
        pdf.set_text_color(0, 0, 0)
        pdf.multi_cell(0, 6,
            "Tautz, J. (2008). The Buzz about Bees: Biology of a Superorganism. Springer.\n"
            "Seeley, T. D. (2010). Honeybee Democracy. Princeton University Press.\n"
            "Free, J. B. (1987). Pheromones of Social Bees. Cornell University Press.\n"
            "Kleinhenz, M. et al. (2003). Hot bees in empty broodnest cells: heating from within. Journal of Experimental Biology, 206(23), 4217-4231.\n"
            "Jones, J. C., & Oldroyd, B. P. (2006). Nest thermoregulation in social insects. Advances in Insect Physiology, 33, 153-191."
        )

        pdf.ln(5)
        pdf.set_font("Arial", "I", 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 10, "Generado automáticamente por HiveSim - 2025", ln=True, align="C")

        # --- Guardar y enviar ---
        output_path = f"informe_{hoy}.pdf"
        pdf.output(output_path)
        return send_file(output_path, as_attachment=True)

    except Exception as e:
        print(f"Error generando PDF: {e}")
        return jsonify({"error": str(e)})

# ---------------- Valores extremos globales ----------------
@app.route("/extremos")
@login_requerido
def extremos():
    try:
        temps, hums = [], []
        for root, _, files in os.walk("data"):
            for f in files:
                if f.endswith(".csv"):
                    ruta = os.path.join(root, f)
                    try:
                        for cod in ["utf-8", "latin1", "iso-8859-1"]:
                            try:
                                df = pd.read_csv(ruta, encoding=cod)
                                break
                            except Exception:
                                continue
                        df = limpiar_columnas(df)
                        temps.extend(df["Temperatura (°C)"].dropna().tolist())
                        hums.extend(df["Humedad (%)"].dropna().tolist())
                    except Exception:
                        continue

        if not temps or not hums:
            return jsonify({"error": "No hay datos para calcular extremos"})

        temp_series = pd.Series(temps)
        hum_series = pd.Series(hums)
        return jsonify({
            "temp_max": float(temp_series.max()),
            "temp_min": float(temp_series.min()),
            "temp_mean": float(temp_series.mean()),
            "temp_std": float(temp_series.std()),
            "hum_max": float(hum_series.max()),
            "hum_min": float(hum_series.min()),
            "hum_mean": float(hum_series.mean()),
            "hum_std": float(hum_series.std())
        })
    except Exception as e:
        print(f"❌ Error procesando extremos: {e}")
        return jsonify({"error": str(e)})

# ---------------- Login / Logout ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        contrasena = request.form["contrasena"]
        conn = sqlite3.connect(DB_USUARIOS)
        c = conn.cursor()
        c.execute("SELECT contrasena FROM usuarios WHERE usuario=?", (usuario,))
        user = c.fetchone()
        conn.close()

        if user and check_password_hash(user[0], contrasena):
            session["usuario"] = usuario
            session.permanent = True
            return redirect(url_for("index"))
        else:
            return render_template("login.html", error="Credenciales incorrectas")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("usuario", None)
    return redirect(url_for("login"))

# ---------------- MONITOREO Y CORREO AUTOMÁTICO ----------------
def enviar_correo(asunto, mensaje):
    try:
        user = os.getenv("EMAIL_USER")
        password = os.getenv("EMAIL_PASS")
        destino = os.getenv("EMAIL_DESTINO")

        msg = EmailMessage()
        msg["From"] = user
        msg["To"] = destino
        msg["Subject"] = asunto
        msg.set_content(mensaje)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(user, password)
            smtp.send_message(msg)
        print(f"📧 Correo enviado: {asunto}")
    except Exception as e:
        print(f"⚠️ Error enviando correo: {e}")

def monitor_sistema():
    while True:
        try:
            ahora = datetime.now()
            diff = (ahora - ultimo_promedio).total_seconds() / 60

            if diff <= 15:
                asunto = "[HiveSim] Estado del sistema apícola"
                mensaje = f"✅ Sistema operativo a las {ahora.strftime('%H:%M:%S')}. Las mediciones continúan correctamente."
            else:
                asunto = "[HiveSim] ⚠️ Alerta de medición"
                mensaje = f"⚠️ No se detectan promedios en los últimos {int(diff)} minutos.\nRevisar la conexión o reiniciar el sistema."

            enviar_correo(asunto, mensaje)
        except Exception as e:
            print(f"Error en monitor_sistema: {e}")

        time.sleep(3600)

threading.Thread(target=monitor_sistema, daemon=True).start()
# ---------------- CLIMA EXTERNO (API OpenWeatherMap) ----------------
@app.route("/clima_externo")
def clima_externo():
    try:
        import requests, os
        API_KEY = os.getenv("OWM_API_KEY")

        if not API_KEY:
            return jsonify({"error": "Falta la clave API"})

        ciudad = "Los Angeles,CL"
        url = f"https://api.openweathermap.org/data/2.5/weather"
        params = {"q": ciudad, "appid": API_KEY, "units": "metric", "lang": "es"}

        r = requests.get(url, params=params, timeout=10)
        if r.status_code != 200:
            return jsonify({"error": f"Error de API ({r.status_code})"})

        data = r.json()
        if "main" not in data:
            return jsonify({"error": "Respuesta sin datos 'main'"})

        temp = data["main"]["temp"]
        hum = data["main"]["humidity"]
        desc = data["weather"][0]["description"].capitalize()

        return jsonify({"temp": temp, "hum": hum, "desc": desc})
    except Exception as e:
        return jsonify({"error": str(e)})


# ---------------- Ejecución ----------------
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))  # Render asigna automáticamente el puerto
    app.run(host="0.0.0.0", port=port, debug=False)

